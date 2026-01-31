from __future__ import annotations

import asyncio
import csv
import json
import os
import re
import sqlite3
import uuid
from io import BytesIO
from datetime import datetime, time as time_value
from urllib.parse import urlparse
from typing import Dict, List, Optional, Tuple

from aiogram import Bot as TelegramBot
from aiogram import Dispatcher
from aiogram.filters import Command
from aiogram.types import (
    CallbackQuery,
    CopyTextButton,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    InputMediaAudio,
    InputMediaDocument,
    InputMediaPhoto,
    InputMediaVideo,
    KeyboardButton,
    Message,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    WebAppInfo,
    FSInputFile,
)
from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

app = FastAPI(title="Bot Builder API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

DB_PATH = os.getenv("BOT_DB", "bot_builder.db")
UPLOAD_DIR = os.getenv("BOT_UPLOADS", "uploads")
FILES_DIR = os.getenv("BOT_FILES", "files")
EXCEL_DIR = os.path.join(FILES_DIR, "excel")
TEXT_DIR = os.path.join(FILES_DIR, "text")
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(EXCEL_DIR, exist_ok=True)
os.makedirs(TEXT_DIR, exist_ok=True)


class Flow(BaseModel):
    nodes: List[dict] = Field(default_factory=list)
    edges: List[dict] = Field(default_factory=list)


class BotCreate(BaseModel):
    name: str = Field(min_length=1)
    token: Optional[str] = None


class BotUpdate(BaseModel):
    name: Optional[str] = Field(default=None, min_length=1)
    token: Optional[str] = None
    status: Optional[str] = None


class Bot(BaseModel):
    id: str
    name: str
    token: Optional[str] = None
    status: str = "stopped"
    flow: Flow = Field(default_factory=Flow)


RUNNING_BOTS: Dict[str, Dict[str, object]] = {}
FLOW_CACHE: Dict[str, Flow] = {}


def get_connection() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    with get_connection() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS bots (
                id TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                token TEXT,
                status TEXT NOT NULL,
                flow TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS user_status (
                bot_id TEXT NOT NULL,
                user_id INTEGER NOT NULL,
                username TEXT,
                first_name TEXT,
                last_name TEXT,
                photo_file_id TEXT,
                status TEXT,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (bot_id, user_id)
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS bot_chats (
                bot_id TEXT NOT NULL,
                chat_id INTEGER NOT NULL,
                title TEXT,
                username TEXT,
                type TEXT,
                is_admin INTEGER DEFAULT 0,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (bot_id, chat_id)
            )
            """
        )
        conn.commit()


init_db()


@app.on_event("startup")
async def resume_running_bots() -> None:
    try:
        with get_connection() as conn:
            rows = conn.execute("SELECT * FROM bots WHERE status = 'running'").fetchall()
    except Exception:
        return
    for row in rows:
        try:
            bot = row_to_bot(row)
        except Exception:
            continue
        if not bot.token:
            continue
        if bot.id in RUNNING_BOTS:
            continue
        FLOW_CACHE[bot.id] = bot.flow
        asyncio.create_task(run_bot_polling(bot))


def row_to_bot(row: sqlite3.Row) -> Bot:
    flow = json.loads(row["flow"]) if row["flow"] else {"nodes": [], "edges": []}
    return Bot(
        id=row["id"],
        name=row["name"],
        token=row["token"],
        status=row["status"],
        flow=Flow(**flow),
    )


def get_user_row(bot_id: str, user_id: int) -> Optional[sqlite3.Row]:
    with get_connection() as conn:
        return conn.execute(
            "SELECT * FROM user_status WHERE bot_id = ? AND user_id = ?",
            (bot_id, user_id),
        ).fetchone()


def upsert_user_row(
    bot_id: str,
    user_id: int,
    username: Optional[str],
    first_name: Optional[str],
    last_name: Optional[str],
    photo_file_id: Optional[str],
    status: Optional[str],
) -> None:
    with get_connection() as conn:
        conn.execute(
            """
            INSERT INTO user_status (
                bot_id,
                user_id,
                username,
                first_name,
                last_name,
                photo_file_id,
                status,
                updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(bot_id, user_id) DO UPDATE SET
                username = excluded.username,
                first_name = excluded.first_name,
                last_name = excluded.last_name,
                photo_file_id = excluded.photo_file_id,
                status = excluded.status,
                updated_at = CURRENT_TIMESTAMP
            """,
            (bot_id, user_id, username, first_name, last_name, photo_file_id, status),
        )
        conn.commit()


def set_user_status(bot_id: str, user_id: int, status: str) -> None:
    value = (status or "").strip()
    with get_connection() as conn:
        existing = conn.execute(
            "SELECT 1 FROM user_status WHERE bot_id = ? AND user_id = ?",
            (bot_id, user_id),
        ).fetchone()
        if existing:
            conn.execute(
                "UPDATE user_status SET status = ?, updated_at = CURRENT_TIMESTAMP WHERE bot_id = ? AND user_id = ?",
                (value, bot_id, user_id),
            )
        else:
            conn.execute(
                """
                INSERT INTO user_status (bot_id, user_id, status, updated_at)
                VALUES (?, ?, ?, CURRENT_TIMESTAMP)
                """,
                (bot_id, user_id, value),
            )
        conn.commit()


def get_user_status(bot_id: str, user_id: int) -> str:
    row = get_user_row(bot_id, user_id)
    if not row:
        return ""
    return (row["status"] or "").strip()


def list_users(bot_id: str) -> List[dict]:
    with get_connection() as conn:
        rows = conn.execute(
            "SELECT user_id, username, first_name, last_name FROM user_status WHERE bot_id = ? ORDER BY updated_at DESC",
            (bot_id,),
        ).fetchall()
    return [
        {
            "id": row["user_id"],
            "username": row["username"],
            "first_name": row["first_name"],
            "last_name": row["last_name"],
        }
        for row in rows
    ]


def list_user_entries(bot_id: str) -> List[dict]:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT user_id, username, first_name, last_name, status
            FROM user_status
            WHERE bot_id = ?
            ORDER BY updated_at DESC
            """,
            (bot_id,),
        ).fetchall()
    return [
        {
            "id": row["user_id"],
            "username": row["username"],
            "first_name": row["first_name"],
            "last_name": row["last_name"],
            "status": row["status"] or "",
        }
        for row in rows
    ]


def upsert_chat_row(
    bot_id: str,
    chat_id: int,
    title: Optional[str],
    username: Optional[str],
    chat_type: Optional[str],
    is_admin: bool,
) -> None:
    with get_connection() as conn:
        conn.execute(
            """
            INSERT INTO bot_chats (
                bot_id,
                chat_id,
                title,
                username,
                type,
                is_admin,
                updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(bot_id, chat_id) DO UPDATE SET
                title = excluded.title,
                username = excluded.username,
                type = excluded.type,
                is_admin = excluded.is_admin,
                updated_at = CURRENT_TIMESTAMP
            """,
            (bot_id, chat_id, title, username, chat_type, 1 if is_admin else 0),
        )
        conn.commit()


def list_chats(bot_id: str) -> List[dict]:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT chat_id, title, username, type
            FROM bot_chats
            WHERE bot_id = ? AND is_admin = 1
            ORDER BY updated_at DESC
            """,
            (bot_id,),
        ).fetchall()
    return [
        {
            "id": row["chat_id"],
            "title": row["title"],
            "username": row["username"],
            "type": row["type"],
        }
        for row in rows
    ]


async def fetch_profile_photo_id(telegram_bot: TelegramBot, user_id: int) -> Optional[str]:
    try:
        photos = await telegram_bot.get_user_profile_photos(user_id, limit=1)
    except Exception:
        return None
    if not photos or not photos.photos:
        return None
    first = photos.photos[0]
    if not first:
        return None
    return first[-1].file_id if first else None


async def ensure_user_row(bot_id: str, user: Optional[object], telegram_bot: Optional[TelegramBot]) -> None:
    if not user:
        return
    user_id = getattr(user, "id", None)
    if user_id is None:
        return
    row = get_user_row(bot_id, user_id)
    existing_status = row["status"] if row else ""
    existing_photo = row["photo_file_id"] if row else None
    photo_file_id = existing_photo
    if photo_file_id is None and telegram_bot:
        photo_file_id = await fetch_profile_photo_id(telegram_bot, user_id)
    if photo_file_id is None:
        photo_file_id = existing_photo or ""
    upsert_user_row(
        bot_id,
        user_id,
        getattr(user, "username", None),
        getattr(user, "first_name", None),
        getattr(user, "last_name", None),
        photo_file_id,
        existing_status,
    )


async def ensure_chat_row(
    bot_id: str,
    chat: Optional[object],
    telegram_bot: TelegramBot,
    bot_user_id: int,
    admin_cache: Dict[int, bool],
) -> None:
    if not chat:
        return
    chat_type = getattr(chat, "type", None)
    if chat_type not in ("group", "supergroup", "channel"):
        return
    chat_id = getattr(chat, "id", None)
    if chat_id is None:
        return
    is_admin = admin_cache.get(chat_id)
    if is_admin is None:
        try:
            member = await telegram_bot.get_chat_member(chat_id, bot_user_id)
            status = getattr(member, "status", "") or ""
            is_admin = status in ("administrator", "creator")
        except Exception:
            is_admin = False
        admin_cache[chat_id] = is_admin
    if not is_admin:
        return
    upsert_chat_row(
        bot_id,
        chat_id,
        getattr(chat, "title", None),
        getattr(chat, "username", None),
        chat_type,
        True,
    )


def get_bot_or_404(bot_id: str) -> Bot:
    with get_connection() as conn:
        row = conn.execute("SELECT * FROM bots WHERE id = ?", (bot_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Bot not found")
    return row_to_bot(row)


def update_bot_row(bot: Bot) -> None:
    with get_connection() as conn:
        conn.execute(
            "UPDATE bots SET name = ?, token = ?, status = ?, flow = ? WHERE id = ?",
            (
                bot.name,
                bot.token,
                bot.status,
                json.dumps(bot.flow.model_dump()),
                bot.id,
            ),
        )
        conn.commit()


def normalize_command(text: str) -> str:
    cleaned = text.strip()
    if not cleaned.startswith("/"):
        return ""
    value = cleaned[1:]
    if not value:
        return ""
    value = value.split()[0]
    value = value.split("@")[0]
    return value.lower()


def sanitize_filename(value: str, fallback: str = "data") -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9_-]+", "_", (value or "").strip())
    cleaned = cleaned.strip("_")
    return cleaned or fallback


def get_excel_path(bot_id: str, name: str) -> str:
    filename = f"{sanitize_filename(name)}.csv"
    return os.path.join(EXCEL_DIR, f"{sanitize_filename(bot_id)}__{filename}")


def get_text_path(bot_id: str, name: str) -> str:
    filename = f"{sanitize_filename(name)}.txt"
    return os.path.join(TEXT_DIR, f"{sanitize_filename(bot_id)}__{filename}")


def render_template(
    text: str,
    message: Message,
    user: Optional[object] = None,
    chat_id_override: Optional[int] = None,
    row_data: Optional[dict] = None,
) -> str:
    if not text:
        return ""
    source = user or message.from_user
    first_name = getattr(source, "first_name", "") if source else ""
    last_name = getattr(source, "last_name", "") if source else ""
    username = getattr(source, "username", "") if source else ""
    incoming_text = (message.text or message.caption or "").strip() if message else ""
    message_id = getattr(message, "message_id", "") if message else ""
    photo_id = ""
    if message and getattr(message, "photo", None):
        try:
            photo_id = message.photo[-1].file_id
        except Exception:
            photo_id = ""
    video_id = getattr(getattr(message, "video", None), "file_id", "") if message else ""
    audio_id = getattr(getattr(message, "audio", None), "file_id", "") if message else ""
    voice_id = getattr(getattr(message, "voice", None), "file_id", "") if message else ""
    document_id = getattr(getattr(message, "document", None), "file_id", "") if message else ""
    sticker_id = getattr(getattr(message, "sticker", None), "file_id", "") if message else ""
    contact_phone = getattr(getattr(message, "contact", None), "phone_number", "") if message else ""
    location_lat = ""
    location_lon = ""
    if message and getattr(message, "location", None):
        try:
            location_lat = str(message.location.latitude)
            location_lon = str(message.location.longitude)
        except Exception:
            location_lat = ""
            location_lon = ""
    if chat_id_override is not None:
        chat_id = chat_id_override
    else:
        chat_id = message.chat.id if message.chat else ""
    replacements = {
        "{text}": incoming_text,
        "{name}": first_name,
        "{first_name}": first_name,
        "{last_name}": last_name,
        "{username}": f"@{username}" if username else "",
        "{chat_id}": str(chat_id),
        "{full_name}": " ".join(part for part in [first_name, last_name] if part),
        "{message_id}": str(message_id),
        "{photo_id}": str(photo_id),
        "{video_id}": str(video_id),
        "{audio_id}": str(audio_id),
        "{voice_id}": str(voice_id),
        "{document_id}": str(document_id),
        "{sticker_id}": str(sticker_id),
        "{contact_phone}": str(contact_phone),
        "{location_lat}": str(location_lat),
        "{location_lon}": str(location_lon),
    }
    result = text
    for key, value in replacements.items():
        result = result.replace(key, value)
    if row_data:
        lower_map = {str(key).lower(): str(value) for key, value in row_data.items() if value is not None}
        def replace_row(match: re.Match) -> str:
            key = match.group(1).strip().lower()
            return lower_map.get(key, "")
        result = re.sub(r"\{row\[([^\]]+)\]\}", replace_row, result)
    return result


def find_command_node(flow: Flow, command: str) -> Optional[dict]:
    command_lower = command.lower()
    for node in flow.nodes:
        data = node.get("data", {})
        if data.get("kind") != "command":
            continue
        command_text = (data.get("commandText") or "/start").strip()
        if not command_text:
            continue
        normalized = command_text.lstrip("/").strip().lower()
        if not normalized:
            continue
        if normalized == command_lower:
            return node
    return None


def parse_timer_seconds(node: dict) -> float:
    data = node.get("data", {})
    raw = data.get("timerSeconds") or 0
    try:
        value = float(raw)
    except (TypeError, ValueError):
        value = 0.0
    return max(0.0, value)


def parse_chat_id(node: dict) -> Optional[int]:
    data = node.get("data", {})
    raw = data.get("chatId")
    try:
        value = int(raw)
    except (TypeError, ValueError):
        return None
    return value


def parse_subscription_chat_id(node: dict) -> Optional[int]:
    data = node.get("data", {})
    raw = data.get("subscriptionChatId")
    try:
        value = int(raw)
    except (TypeError, ValueError):
        return None
    return value


async def is_user_subscribed(telegram_bot: TelegramBot, chat_id: int, user_id: int) -> bool:
    try:
        member = await telegram_bot.get_chat_member(chat_id, user_id)
    except Exception:
        return False
    status = getattr(member, "status", "") or ""
    if status in ("creator", "administrator", "member"):
        return True
    if status == "restricted":
        return bool(getattr(member, "is_member", False))
    return False


def match_condition(message: Message, node: dict, bot_id: Optional[str] = None, user_id: Optional[int] = None) -> bool:
    data = node.get("data", {})
    text_value = (message.text or message.caption or "").strip()
    checks = []
    condition_type = (data.get("conditionType") or "").strip()
    condition_text = (data.get("conditionText") or "").strip()

    if condition_type == "text":
        if not condition_text:
            return False
        checks.append(text_value.lower() == condition_text.lower())
    elif condition_type == "text_contains":
        if not condition_text:
            return False
        checks.append(condition_text.lower() in text_value.lower())
    elif condition_type == "status":
        if not condition_text or not bot_id or user_id is None:
            return False
        status_value = get_user_status(bot_id, user_id)
        checks.append(status_value.lower() == condition_text.lower())
    elif condition_type == "has_text":
        checks.append(bool(text_value))
    elif condition_type == "has_number":
        checks.append(bool(re.search(r"\d", text_value)))
    elif condition_type == "has_photo":
        checks.append(bool(message.photo))
    elif condition_type == "has_video":
        checks.append(message.video is not None)
    elif condition_type == "has_audio":
        checks.append(message.audio is not None)
    elif condition_type == "has_voice":
        checks.append(message.voice is not None)
    elif condition_type == "has_document":
        checks.append(message.document is not None)
    elif condition_type == "has_sticker":
        checks.append(message.sticker is not None)
    elif condition_type == "has_contact":
        checks.append(message.contact is not None)
    elif condition_type == "has_location":
        checks.append(message.location is not None)
    elif condition_text:
        checks.append(text_value.lower() == condition_text.lower())
    else:
        if data.get("conditionHasText"):
            checks.append(bool(text_value))
        if data.get("conditionHasNumber"):
            checks.append(bool(re.search(r"\d", text_value)))
        if data.get("conditionHasPhoto"):
            checks.append(bool(message.photo))
        if data.get("conditionHasVideo"):
            checks.append(message.video is not None)
        if data.get("conditionHasAudio"):
            checks.append(message.audio is not None)
        if data.get("conditionHasLocation"):
            checks.append(message.location is not None)

    length_op = data.get("conditionLengthOp")
    length_value = data.get("conditionLengthValue")
    if length_op and length_value is not None and condition_type in ("has_text", "has_number"):
        try:
            length_value = int(length_value)
            length_actual = len(text_value)
            if length_op == "lt":
                checks.append(length_actual < length_value)
            elif length_op == "lte":
                checks.append(length_actual <= length_value)
            elif length_op == "eq":
                checks.append(length_actual == length_value)
            elif length_op == "gte":
                checks.append(length_actual >= length_value)
            elif length_op == "gt":
                checks.append(length_actual > length_value)
        except (TypeError, ValueError):
            pass

    if not checks:
        return False
    return all(checks)


def extract_record_value(message: Message, record_field: str) -> str:
    field = (record_field or "").strip()
    if field == "text":
        return (message.text or message.caption or "").strip()
    if field == "message_id":
        return str(getattr(message, "message_id", ""))
    if field == "name":
        return getattr(message.from_user, "first_name", "") if message.from_user else ""
    if field == "first_name":
        return getattr(message.from_user, "first_name", "") if message.from_user else ""
    if field == "last_name":
        return getattr(message.from_user, "last_name", "") if message.from_user else ""
    if field == "username":
        username = getattr(message.from_user, "username", "") if message.from_user else ""
        return f"@{username}" if username else ""
    if field == "full_name":
        first = getattr(message.from_user, "first_name", "") if message.from_user else ""
        last = getattr(message.from_user, "last_name", "") if message.from_user else ""
        return " ".join(part for part in [first, last] if part)
    if field == "chat_id":
        return str(message.chat.id) if message.chat else ""
    if field == "contact_phone":
        return getattr(getattr(message, "contact", None), "phone_number", "") or ""
    if field == "location_lat":
        return str(getattr(getattr(message, "location", None), "latitude", "") or "")
    if field == "location_lon":
        return str(getattr(getattr(message, "location", None), "longitude", "") or "")
    if field == "photo_id":
        if message.photo:
            return message.photo[-1].file_id
        return ""
    if field == "video_id":
        return message.video.file_id if message.video else ""
    if field == "audio_id":
        return message.audio.file_id if message.audio else ""
    if field == "voice_id":
        return message.voice.file_id if message.voice else ""
    if field == "document_id":
        return message.document.file_id if message.document else ""
    if field == "sticker_id":
        return message.sticker.file_id if message.sticker else ""
    return ""


def extract_record_value_from_entry(entry: dict, record_field: str) -> str:
    field = (record_field or "").strip()
    if field == "name":
        return entry.get("first_name", "") or ""
    if field == "first_name":
        return entry.get("first_name", "") or ""
    if field == "last_name":
        return entry.get("last_name", "") or ""
    if field == "username":
        username = entry.get("username") or ""
        return f"@{username}" if username else ""
    if field == "full_name":
        first = entry.get("first_name") or ""
        last = entry.get("last_name") or ""
        return " ".join(part for part in [first, last] if part)
    if field == "chat_id":
        return str(entry.get("id", ""))
    return ""


def append_to_text_file(bot_id: str, file_name: str, value: str) -> None:
    path = get_text_path(bot_id, file_name)
    line = (value or "").strip()
    if not line:
        return
    with open(path, "a", encoding="utf-8") as file_obj:
        file_obj.write(line + "\n")


def append_to_excel_file(bot_id: str, file_name: str, column: str, value: str) -> None:
    path = get_excel_path(bot_id, file_name)
    column = (column or "").strip() or "Value"
    value = (value or "").strip()
    rows: List[dict] = []
    headers: List[str] = []
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as file_obj:
            reader = csv.DictReader(file_obj)
            headers = reader.fieldnames or []
            rows = list(reader)
    if column not in headers:
        headers.append(column)
    new_row = {header: "" for header in headers}
    new_row[column] = value
    rows.append(new_row)
    with open(path, "w", encoding="utf-8", newline="") as file_obj:
        writer = csv.DictWriter(file_obj, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def resolve_excel_file_info(flow: Flow, nodes_by_id: dict, column_id: str) -> Optional[dict]:
    for edge in flow.edges:
        if edge.get("target") != column_id:
            continue
        source_node = nodes_by_id.get(edge.get("source"))
        if not source_node:
            continue
        if source_node.get("data", {}).get("kind") != "excel_file":
            continue
        file_name = source_node.get("data", {}).get("fileName") or "data"
        return {"type": "excel", "name": file_name}
    return None


async def collect_content_targets_with_delay(
    flow: Flow,
    source_id: str,
    message: Message | None = None,
    bot_id: Optional[str] = None,
    user_id: Optional[int] = None,
) -> List[Tuple[dict, float, Optional[int], Optional[dict]]]:
    nodes_by_id = {node.get("id"): node for node in flow.nodes}
    results: List[Tuple[dict, float, Optional[int], Optional[dict]]] = []
    seen_targets: set[tuple[str, Optional[int]]] = set()
    visited: set[tuple[str, Optional[int]]] = set()
    queue: List[Tuple[str, float, Optional[int], Optional[str], Optional[dict], Optional[str], Optional[dict]]] = [
        (source_id, 0.0, None, None, None, None, None)
    ]
    effective_user_id = user_id
    if effective_user_id is None and message and message.from_user:
        effective_user_id = message.from_user.id
    while queue:
        current_id, delay, target_chat_id, record_value, file_info, column_name, row_data = queue.pop(0)
        if not current_id or (current_id, target_chat_id) in visited:
            continue
        visited.add((current_id, target_chat_id))
        current_node = nodes_by_id.get(current_id)
        current_kind = current_node.get("data", {}).get("kind") if current_node else None
        if current_kind == "record" and message:
            record_field = current_node.get("data", {}).get("recordField") or ""
            record_value = extract_record_value(message, record_field)
        if current_kind in ("excel_file", "text_file") and current_node:
            file_name = current_node.get("data", {}).get("fileName") or "data"
            file_type = "excel" if current_kind == "excel_file" else "text"
            file_info = {"type": file_type, "name": file_name}
            if file_type == "text" and record_value is not None and bot_id:
                append_to_text_file(bot_id, file_name, record_value)
        if current_kind == "excel_column" and current_node:
            column_name = current_node.get("data", {}).get("columnName") or "Value"
            if file_info is None:
                file_info = resolve_excel_file_info(flow, nodes_by_id, current_id)
            if file_info and file_info.get("type") == "excel" and bot_id:
                append_to_excel_file(bot_id, file_info.get("name") or "data", column_name, record_value or "")
        if current_kind == "chat" and current_node:
            override_chat_id = parse_chat_id(current_node)
            if override_chat_id is not None:
                target_chat_id = override_chat_id
        condition_pass = None
        if current_kind == "condition" and message:
            condition_pass = match_condition(message, current_node, bot_id, effective_user_id)
        subscription_pass = None
        if current_kind == "subscription" and message and effective_user_id is not None:
            chat_id = parse_subscription_chat_id(current_node)
            if chat_id is not None:
                subscription_pass = await is_user_subscribed(message.bot, chat_id, effective_user_id)
        search_pass = None
        if current_kind == "file_search":
            search_pass = False
            row_data = None
            payload = current_node.get("data", {})
            search_source = (payload.get("searchSource") or "incoming").strip()
            manual_value = (payload.get("searchValue") or "").strip()
            search_value = ""
            if search_source == "manual" and manual_value:
                if message:
                    search_value = render_template(manual_value, message, message.from_user)
                else:
                    search_value = manual_value
            else:
                search_value = (record_value or "").strip()
                if not search_value and message:
                    search_value = (message.text or message.caption or "").strip()
            search_column = (payload.get("searchColumnName") or "").strip()
            resolved_file_info = file_info
            resolved_column = search_column or column_name or ""
            for edge in flow.edges:
                if edge.get("target") != current_id:
                    continue
                source_node = nodes_by_id.get(edge.get("source"))
                if not source_node:
                    continue
                source_kind = source_node.get("data", {}).get("kind")
                if source_kind == "text_file":
                    if resolved_file_info is None:
                        file_name = source_node.get("data", {}).get("fileName") or "data"
                        resolved_file_info = {"type": "text", "name": file_name}
                elif source_kind == "excel_column":
                    if not search_column:
                        resolved_column = (source_node.get("data", {}).get("columnName") or "").strip() or resolved_column
                    if resolved_file_info is None:
                        resolved_file_info = resolve_excel_file_info(flow, nodes_by_id, source_node.get("id") or "")
                elif source_kind == "excel_file":
                    if resolved_file_info is None:
                        file_name = source_node.get("data", {}).get("fileName") or "data"
                        resolved_file_info = {"type": "excel", "name": file_name}
            for edge in flow.edges:
                if edge.get("source") != current_id:
                    continue
                target_node = nodes_by_id.get(edge.get("target"))
                if not target_node or target_node.get("data", {}).get("kind") != "excel_column":
                    continue
                if not search_column:
                    resolved_column = (target_node.get("data", {}).get("columnName") or "").strip() or resolved_column
                if resolved_file_info is None:
                    resolved_file_info = resolve_excel_file_info(flow, nodes_by_id, target_node.get("id") or "")
                if resolved_column and resolved_file_info:
                    break
            if not resolved_column or resolved_file_info is None:
                for edge in flow.edges:
                    if edge.get("target") != current_id:
                        continue
                    source_node = nodes_by_id.get(edge.get("source"))
                    if not source_node or source_node.get("data", {}).get("kind") != "excel_column":
                        continue
                    if not search_column:
                        resolved_column = (source_node.get("data", {}).get("columnName") or "").strip() or resolved_column
                    if resolved_file_info is None:
                        resolved_file_info = resolve_excel_file_info(flow, nodes_by_id, source_node.get("id") or "")
                    if resolved_column and resolved_file_info:
                        break
            if not resolved_column or resolved_file_info is None:
                for edge in flow.edges:
                    if edge.get("target") != current_id:
                        continue
                    source_node = nodes_by_id.get(edge.get("source"))
                    if not source_node or source_node.get("data", {}).get("kind") != "excel_column":
                        continue
                    if not search_column:
                        resolved_column = (source_node.get("data", {}).get("columnName") or "").strip() or resolved_column
                    if resolved_file_info is None:
                        resolved_file_info = resolve_excel_file_info(flow, nodes_by_id, source_node.get("id") or "")
                    if resolved_column and resolved_file_info:
                        break
            if resolved_file_info and search_value:
                if resolved_file_info.get("type") == "excel" and resolved_column and bot_id:
                    path = get_excel_path(bot_id, resolved_file_info.get("name") or "data")
                    if os.path.exists(path):
                        with open(path, "r", encoding="utf-8") as file_obj:
                            reader = csv.DictReader(file_obj)
                            for row in reader:
                                cell_value = (row.get(resolved_column) or "").strip()
                                if cell_value.lower() == search_value.lower():
                                    row_data = row
                                    search_pass = True
                                    break
                if resolved_file_info.get("type") == "text" and bot_id:
                    path = get_text_path(bot_id, resolved_file_info.get("name") or "data")
                    if os.path.exists(path):
                        with open(path, "r", encoding="utf-8") as file_obj:
                            for line in file_obj:
                                if line.strip().lower() == search_value.lower():
                                    row_data = {"value": line.strip()}
                                    search_pass = True
                                    break
            file_info = resolved_file_info or file_info
            column_name = resolved_column or column_name
        for edge in flow.edges:
            if edge.get("source") != current_id:
                continue
            if current_kind == "file_search":
                target_node = nodes_by_id.get(edge.get("target"))
                if target_node and target_node.get("data", {}).get("kind") == "excel_column":
                    continue
            if current_kind in ("condition", "subscription", "file_search"):
                handle = edge.get("sourceHandle") or "true"
                if current_kind == "condition":
                    result = condition_pass
                elif current_kind == "subscription":
                    result = subscription_pass
                else:
                    result = search_pass
                expected = "true" if result else "false"
                if handle != expected:
                    continue
            target_node = nodes_by_id.get(edge.get("target"))
            if not target_node:
                continue
            kind = target_node.get("data", {}).get("kind")
            if kind in ("message", "image", "video", "audio", "document", "delete_message", "edit_message"):
                target_id = target_node.get("id") or ""
                key = (target_id, target_chat_id)
                if target_id and key not in seen_targets:
                    seen_targets.add(key)
                    results.append((target_node, delay, target_chat_id, row_data))
            elif kind == "timer":
                next_delay = delay + parse_timer_seconds(target_node)
                queue.append((target_node.get("id") or "", next_delay, target_chat_id, record_value, file_info, column_name, row_data))
            elif kind in ("condition", "subscription", "record", "excel_file", "text_file", "excel_column", "file_search", "broadcast", "task"):
                queue.append((target_node.get("id") or "", delay, target_chat_id, record_value, file_info, column_name, row_data))
            elif kind == "status_set":
                if bot_id and effective_user_id is not None:
                    set_user_status(
                        bot_id,
                        effective_user_id,
                        target_node.get("data", {}).get("statusValue") or "",
                    )
                queue.append((target_node.get("id") or "", delay, target_chat_id, record_value, file_info, column_name, row_data))
            elif kind == "status_get":
                queue.append((target_node.get("id") or "", delay, target_chat_id, record_value, file_info, column_name, row_data))
            elif kind == "chat":
                queue.append((target_node.get("id") or "", delay, target_chat_id, record_value, file_info, column_name, row_data))
    return results


def parse_task_interval_minutes(node: dict) -> int:
    data = node.get("data", {})
    raw = data.get("taskIntervalMinutes")
    try:
        value = int(raw)
    except (TypeError, ValueError):
        value = 60
    if value <= 0:
        return 60
    return value


def parse_task_daily_time(node: dict) -> Optional[time_value]:
    data = node.get("data", {})
    raw = (data.get("taskDailyTime") or "").strip()
    if not raw:
        return None
    try:
        parts = raw.split(":")
        hour = int(parts[0])
        minute = int(parts[1]) if len(parts) > 1 else 0
        return time_value(hour=hour, minute=minute)
    except (ValueError, IndexError):
        return None


def parse_task_run_at(node: dict) -> Optional[datetime]:
    data = node.get("data", {})
    raw = (data.get("taskRunAt") or "").strip()
    if not raw:
        return None
    for fmt in ("%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    return None


def should_run_task(node: dict, now: datetime, last_run: Optional[datetime]) -> bool:
    schedule_type = (node.get("data", {}).get("taskScheduleType") or "interval").strip()
    if schedule_type == "daily":
        daily_time = parse_task_daily_time(node)
        if daily_time is None:
            daily_time = time_value(hour=10, minute=0)
        if now.time() < daily_time:
            return False
        if last_run is None:
            return True
        return last_run.date() < now.date()
    if schedule_type == "datetime":
        run_at = parse_task_run_at(node)
        if not run_at:
            return False
        if last_run is not None:
            return False
        return now >= run_at
    interval_minutes = parse_task_interval_minutes(node)
    if last_run is None:
        return True
    elapsed = (now - last_run).total_seconds()
    return elapsed >= interval_minutes * 60


def match_condition_for_entry(node: dict, entry: dict) -> bool:
    data = node.get("data", {})
    condition_type = (data.get("conditionType") or "").strip()
    condition_text = (data.get("conditionText") or "").strip()
    if condition_type == "status":
        if not condition_text:
            return False
        status_value = (entry.get("status") or "").strip()
        return status_value.lower() == condition_text.lower()
    return False


async def collect_scheduled_targets(
    flow: Flow,
    source_id: str,
    bot_id: str,
    telegram_bot: TelegramBot,
) -> List[Tuple[dict, float, int, Optional[dict], Optional[dict]]]:
    nodes_by_id = {node.get("id"): node for node in flow.nodes}
    results: List[Tuple[dict, float, int, Optional[dict], Optional[dict]]] = []
    user_entries = list_user_entries(bot_id)
    user_by_id = {entry["id"]: entry for entry in user_entries}
    visited: set[tuple[str, Optional[int], Optional[int]]] = set()
    queue: List[Tuple[str, float, Optional[int], Optional[dict], Optional[str], Optional[dict], Optional[str], Optional[dict]]] = [
        (source_id, 0.0, None, None, None, None, None, None)
    ]
    while queue:
        current_id, delay, target_chat_id, entry, record_value, file_info, column_name, row_data = queue.pop(0)
        entry_id = entry["id"] if entry else None
        if not current_id or (current_id, target_chat_id, entry_id) in visited:
            continue
        visited.add((current_id, target_chat_id, entry_id))
        current_node = nodes_by_id.get(current_id)
        current_kind = current_node.get("data", {}).get("kind") if current_node else None

        if current_kind == "broadcast" and current_node:
            for user_entry in user_entries:
                for edge in flow.edges:
                    if edge.get("source") != current_id:
                        continue
                    queue.append((edge.get("target") or "", delay, user_entry["id"], user_entry, None, file_info, column_name, row_data))
            continue

        if current_kind == "record" and entry:
            record_field = current_node.get("data", {}).get("recordField") or ""
            record_value = extract_record_value_from_entry(entry, record_field)
        if current_kind in ("excel_file", "text_file") and current_node:
            file_name = current_node.get("data", {}).get("fileName") or "data"
            file_type = "excel" if current_kind == "excel_file" else "text"
            file_info = {"type": file_type, "name": file_name}
            if file_type == "text" and record_value is not None:
                append_to_text_file(bot_id, file_name, record_value)
        if current_kind == "excel_column" and current_node:
            column_name = current_node.get("data", {}).get("columnName") or "Value"
            if file_info is None:
                file_info = resolve_excel_file_info(flow, nodes_by_id, current_id)
            if file_info and file_info.get("type") == "excel":
                append_to_excel_file(bot_id, file_info.get("name") or "data", column_name, record_value or "")

        if current_kind == "chat" and current_node:
            override_chat_id = parse_chat_id(current_node)
            if override_chat_id is not None:
                target_chat_id = override_chat_id
                entry = user_by_id.get(override_chat_id, entry)
        condition_pass = None
        if current_kind == "condition" and entry:
            condition_pass = match_condition_for_entry(current_node, entry)
        subscription_pass = None
        if current_kind == "subscription" and entry:
            chat_id = parse_subscription_chat_id(current_node)
            if chat_id is not None:
                subscription_pass = await is_user_subscribed(telegram_bot, chat_id, entry["id"])
        search_pass = None
        if current_kind == "file_search":
            search_pass = False
            row_data = None
            payload = current_node.get("data", {})
            search_source = (payload.get("searchSource") or "incoming").strip()
            manual_value = (payload.get("searchValue") or "").strip()
            search_value = ""
            if search_source == "manual" and manual_value:
                fake_message = type("Obj", (), {"chat": type("Obj", (), {"id": target_chat_id or 0})()})()
                user_obj = None
                if entry:
                    user_obj = type(
                        "Obj",
                        (),
                        {
                            "id": entry.get("id"),
                            "username": entry.get("username"),
                            "first_name": entry.get("first_name"),
                            "last_name": entry.get("last_name"),
                        },
                    )()
                search_value = render_template(manual_value, fake_message, user_obj, target_chat_id or 0)
            else:
                search_value = (record_value or "").strip()
            search_column = (payload.get("searchColumnName") or "").strip()
            resolved_file_info = file_info
            resolved_column = search_column or column_name or ""
            for edge in flow.edges:
                if edge.get("target") != current_id:
                    continue
                source_node = nodes_by_id.get(edge.get("source"))
                if not source_node:
                    continue
                source_kind = source_node.get("data", {}).get("kind")
                if source_kind == "text_file":
                    if resolved_file_info is None:
                        file_name = source_node.get("data", {}).get("fileName") or "data"
                        resolved_file_info = {"type": "text", "name": file_name}
                elif source_kind == "excel_column":
                    if not search_column:
                        resolved_column = (source_node.get("data", {}).get("columnName") or "").strip() or resolved_column
                    if resolved_file_info is None:
                        resolved_file_info = resolve_excel_file_info(flow, nodes_by_id, source_node.get("id") or "")
                elif source_kind == "excel_file":
                    if resolved_file_info is None:
                        file_name = source_node.get("data", {}).get("fileName") or "data"
                        resolved_file_info = {"type": "excel", "name": file_name}
            for edge in flow.edges:
                if edge.get("source") != current_id:
                    continue
                target_node = nodes_by_id.get(edge.get("target"))
                if not target_node or target_node.get("data", {}).get("kind") != "excel_column":
                    continue
                if not search_column:
                    resolved_column = (target_node.get("data", {}).get("columnName") or "").strip() or resolved_column
                if resolved_file_info is None:
                    resolved_file_info = resolve_excel_file_info(flow, nodes_by_id, target_node.get("id") or "")
                if resolved_column and resolved_file_info:
                    break
            if resolved_file_info and search_value:
                if resolved_file_info.get("type") == "excel" and resolved_column:
                    path = get_excel_path(bot_id, resolved_file_info.get("name") or "data")
                    if os.path.exists(path):
                        with open(path, "r", encoding="utf-8") as file_obj:
                            reader = csv.DictReader(file_obj)
                            for row in reader:
                                cell_value = (row.get(resolved_column) or "").strip()
                                if cell_value.lower() == search_value.lower():
                                    row_data = row
                                    search_pass = True
                                    break
                if resolved_file_info.get("type") == "text":
                    path = get_text_path(bot_id, resolved_file_info.get("name") or "data")
                    if os.path.exists(path):
                        with open(path, "r", encoding="utf-8") as file_obj:
                            for line in file_obj:
                                if line.strip().lower() == search_value.lower():
                                    row_data = {"value": line.strip()}
                                    search_pass = True
                                    break
            file_info = resolved_file_info or file_info
            column_name = resolved_column or column_name

        for edge in flow.edges:
            if edge.get("source") != current_id:
                continue
            if current_kind == "file_search":
                target_node = nodes_by_id.get(edge.get("target"))
                if target_node and target_node.get("data", {}).get("kind") == "excel_column":
                    continue
            if current_kind in ("condition", "subscription", "file_search"):
                handle = edge.get("sourceHandle") or "true"
                if current_kind == "condition":
                    result = condition_pass
                elif current_kind == "subscription":
                    result = subscription_pass
                else:
                    result = search_pass
                expected = "true" if result else "false"
                if handle != expected:
                    continue
            target_node = nodes_by_id.get(edge.get("target"))
            if not target_node:
                continue
            kind = target_node.get("data", {}).get("kind")
            if kind in ("message", "image", "video", "audio", "document", "delete_message", "edit_message"):
                chat_id = target_chat_id or (entry["id"] if entry else None)
                if chat_id is not None:
                    results.append((target_node, delay, chat_id, entry, row_data))
            elif kind == "timer":
                next_delay = delay + parse_timer_seconds(target_node)
                queue.append((target_node.get("id") or "", next_delay, target_chat_id, entry, record_value, file_info, column_name, row_data))
            elif kind in ("condition", "subscription", "broadcast", "record", "excel_file", "text_file", "excel_column", "file_search", "task"):
                queue.append((target_node.get("id") or "", delay, target_chat_id, entry, record_value, file_info, column_name, row_data))
            elif kind == "status_set":
                if entry:
                    set_user_status(bot_id, entry["id"], target_node.get("data", {}).get("statusValue") or "")
                queue.append((target_node.get("id") or "", delay, target_chat_id, entry, record_value, file_info, column_name, row_data))
            elif kind == "status_get":
                queue.append((target_node.get("id") or "", delay, target_chat_id, entry, record_value, file_info, column_name, row_data))
            elif kind == "chat":
                queue.append((target_node.get("id") or "", delay, target_chat_id, entry, record_value, file_info, column_name, row_data))
    return results


async def send_images_via_bot(
    telegram_bot: TelegramBot,
    chat_id: int,
    urls: List[str],
    caption: str = "",
    reply_markup=None,
) -> bool:
    sources = [resolve_upload_source(url) for url in urls]
    sources = [source for source in sources if source]
    if not sources:
        return False
    if len(sources) == 1:
        await telegram_bot.send_photo(chat_id, sources[0], caption=caption or None, reply_markup=reply_markup)
        return True
    media = []
    for idx, source in enumerate(sources):
        if idx == 0 and caption:
            media.append(InputMediaPhoto(media=source, caption=caption))
        else:
            media.append(InputMediaPhoto(media=source))
    await telegram_bot.send_media_group(chat_id, media)
    return True


async def send_videos_via_bot(
    telegram_bot: TelegramBot,
    chat_id: int,
    urls: List[str],
    caption: str = "",
    reply_markup=None,
) -> bool:
    sources = [resolve_upload_source(url) for url in urls]
    sources = [source for source in sources if source]
    if not sources:
        return False
    if len(sources) == 1:
        await telegram_bot.send_video(chat_id, sources[0], caption=caption or None, reply_markup=reply_markup)
        return True
    media = []
    for idx, source in enumerate(sources):
        if idx == 0 and caption:
            media.append(InputMediaVideo(media=source, caption=caption))
        else:
            media.append(InputMediaVideo(media=source))
    await telegram_bot.send_media_group(chat_id, media)
    return True


async def send_audios_via_bot(
    telegram_bot: TelegramBot,
    chat_id: int,
    urls: List[str],
    caption: str = "",
    reply_markup=None,
) -> bool:
    sources = [resolve_upload_source(url) for url in urls]
    sources = [source for source in sources if source]
    if not sources:
        return False
    if len(sources) == 1:
        await telegram_bot.send_audio(chat_id, sources[0], caption=caption or None, reply_markup=reply_markup)
        return True
    media = []
    for idx, source in enumerate(sources):
        if idx == 0 and caption:
            media.append(InputMediaAudio(media=source, caption=caption))
        else:
            media.append(InputMediaAudio(media=source))
    await telegram_bot.send_media_group(chat_id, media)
    return True


async def send_documents_via_bot(
    telegram_bot: TelegramBot,
    chat_id: int,
    urls: List[str],
    caption: str = "",
    reply_markup=None,
) -> bool:
    sources = [resolve_upload_source(url) for url in urls]
    sources = [source for source in sources if source]
    if not sources:
        return False
    if len(sources) == 1:
        await telegram_bot.send_document(chat_id, sources[0], caption=caption or None, reply_markup=reply_markup)
        return True
    media = []
    for idx, source in enumerate(sources):
        if idx == 0 and caption:
            media.append(InputMediaDocument(media=source, caption=caption))
        else:
            media.append(InputMediaDocument(media=source))
    await telegram_bot.send_media_group(chat_id, media)
    return True


async def send_content_node_to_chat(
    flow: Flow,
    telegram_bot: TelegramBot,
    chat_id: int,
    target_node: dict,
    entry: Optional[dict],
    row_data: Optional[dict] = None,
) -> None:
    payload = target_node.get("data", {})
    kind = payload.get("kind")
    if kind == "delete_message":
        return
    is_edit = kind == "edit_message"
    if kind == "message" or is_edit:
        raw_text = payload.get("editMessageText") if is_edit else payload.get("messageText")
        fake_message = type("Obj", (), {"chat": type("Obj", (), {"id": chat_id})()})()
        user_obj = None
        if entry:
            user_obj = type(
                "Obj",
                (),
                {
                    "id": entry.get("id"),
                    "username": entry.get("username"),
                    "first_name": entry.get("first_name"),
                    "last_name": entry.get("last_name"),
                },
            )()
        message_text = render_template((raw_text or "").strip(), fake_message, user_obj, chat_id, row_data=row_data)
        image_urls = collect_image_urls(flow, target_node.get("id") or "")
        video_urls = collect_video_urls(flow, target_node.get("id") or "")
        audio_urls = collect_audio_urls(flow, target_node.get("id") or "")
        document_urls = collect_document_urls(flow, target_node.get("id") or "")
        reply_markup = build_reply_markup(flow, target_node.get("id") or "")
        caption_used = False
        reply_used = False
        if image_urls:
            caption = message_text if message_text and not caption_used else ""
            await send_images_via_bot(telegram_bot, chat_id, image_urls, caption=caption, reply_markup=reply_markup if not reply_used else None)
            if caption:
                caption_used = True
                reply_used = True
        if video_urls:
            caption = message_text if message_text and not caption_used else ""
            await send_videos_via_bot(telegram_bot, chat_id, video_urls, caption=caption, reply_markup=reply_markup if not reply_used else None)
            if caption:
                caption_used = True
                reply_used = True
        if audio_urls:
            caption = message_text if message_text and not caption_used else ""
            await send_audios_via_bot(telegram_bot, chat_id, audio_urls, caption=caption, reply_markup=reply_markup if not reply_used else None)
            if caption:
                caption_used = True
                reply_used = True
        if document_urls:
            caption = message_text if message_text and not caption_used else ""
            await send_documents_via_bot(telegram_bot, chat_id, document_urls, caption=caption, reply_markup=reply_markup if not reply_used else None)
            if caption:
                caption_used = True
                reply_used = True
        if not (image_urls or video_urls or audio_urls or document_urls):
            if message_text:
                await telegram_bot.send_message(chat_id, message_text, reply_markup=reply_markup)
        return
    if kind == "image":
        urls = payload.get("imageUrls") or []
        reply_markup = build_reply_markup(flow, target_node.get("id") or "")
        await send_images_via_bot(telegram_bot, chat_id, urls, reply_markup=reply_markup)
        return
    if kind == "video":
        urls = payload.get("videoUrls") or []
        reply_markup = build_reply_markup(flow, target_node.get("id") or "")
        await send_videos_via_bot(telegram_bot, chat_id, urls, reply_markup=reply_markup)
        return
    if kind == "audio":
        urls = payload.get("audioUrls") or []
        reply_markup = build_reply_markup(flow, target_node.get("id") or "")
        await send_audios_via_bot(telegram_bot, chat_id, urls, reply_markup=reply_markup)
        return
    if kind == "document":
        urls = payload.get("documentUrls") or []
        reply_markup = build_reply_markup(flow, target_node.get("id") or "")
        await send_documents_via_bot(telegram_bot, chat_id, urls, reply_markup=reply_markup)


async def send_scheduled_targets_with_delay(
    flow: Flow,
    telegram_bot: TelegramBot,
    targets: List[Tuple[dict, float, int, Optional[dict], Optional[dict]]],
) -> None:
    if not targets:
        return
    ordered = sorted(targets, key=lambda item: item[1])
    elapsed = 0.0
    for target_node, delay, chat_id, entry, row_data in ordered:
        wait_time = delay - elapsed
        if wait_time > 0:
            await asyncio.sleep(wait_time)
            elapsed = delay
        await send_content_node_to_chat(flow, telegram_bot, chat_id, target_node, entry, row_data)


async def task_scheduler_loop(
    bot: Bot,
    telegram_bot: TelegramBot,
    stop_event: asyncio.Event,
    last_run_by_task: Dict[str, datetime],
) -> None:
    while not stop_event.is_set():
        flow = FLOW_CACHE.get(bot.id) or bot.flow
        now = datetime.now()
        task_nodes = [node for node in flow.nodes if node.get("data", {}).get("kind") == "task"]
        for task_node in task_nodes:
            task_id = task_node.get("id") or ""
            if not task_id:
                continue
            last_run = last_run_by_task.get(task_id)
            if not should_run_task(task_node, now, last_run):
                continue
            targets = await collect_scheduled_targets(flow, task_id, bot.id, telegram_bot)
            await send_scheduled_targets_with_delay(flow, telegram_bot, targets)
            last_run_by_task[task_id] = now
        try:
            await asyncio.wait_for(stop_event.wait(), timeout=20)
        except asyncio.TimeoutError:
            continue
def find_reply_button_by_text(flow: Flow, text: str) -> Optional[dict]:
    needle = text.strip().lower()
    if not needle:
        return None
    for node in flow.nodes:
        data = node.get("data", {})
        if data.get("kind") != "reply_button":
            continue
        label = (data.get("buttonText") or data.get("label") or "").strip().lower()
        if label == needle:
            return node
    return None


def collect_button_rows(flow: Flow, content_node_id: str) -> Tuple[List[List[dict]], List[List[dict]]]:
    nodes_by_id = {node.get("id"): node for node in flow.nodes}
    row_edges = [edge for edge in flow.edges if edge.get("source") == content_node_id]
    row_nodes = []
    direct_buttons = []
    for edge in row_edges:
        target_node = nodes_by_id.get(edge.get("target"))
        if not target_node:
            continue
        kind = target_node.get("data", {}).get("kind")
        if kind == "button_row":
            row_nodes.append(target_node)
        elif kind in ("message_button", "reply_button"):
            direct_buttons.append(target_node)

    inline_rows: List[List[dict]] = []
    reply_rows: List[List[dict]] = []

    if row_nodes:
        for row_node in row_nodes:
            row_buttons = [
                nodes_by_id.get(edge.get("target"))
                for edge in flow.edges
                if edge.get("source") == row_node.get("id")
            ]
            row_buttons = [btn for btn in row_buttons if btn]
            inline = [btn for btn in row_buttons if btn.get("data", {}).get("kind") == "message_button"]
            reply = [btn for btn in row_buttons if btn.get("data", {}).get("kind") == "reply_button"]
            if inline:
                inline_rows.append(inline)
            if reply:
                reply_rows.append(reply)
    else:
        for btn in direct_buttons:
            kind = btn.get("data", {}).get("kind")
            if kind == "message_button":
                inline_rows.append([btn])
            elif kind == "reply_button":
                reply_rows.append([btn])

    return inline_rows, reply_rows


def build_reply_markup(flow: Flow, content_node_id: str):
    nodes_by_id = {node.get("id"): node for node in flow.nodes}
    inline_rows, reply_rows = collect_button_rows(flow, content_node_id)
    has_clear = any(
        nodes_by_id.get(edge.get("target"), {}).get("data", {}).get("kind") == "reply_clear"
        for edge in flow.edges
        if edge.get("source") == content_node_id
    )

    def build_inline_button(btn: dict) -> Optional[InlineKeyboardButton]:
        data = btn.get("data", {})
        text = (data.get("buttonText") or data.get("label") or "").strip()
        if not text:
            return None
        action = (data.get("buttonAction") or "callback").strip()
        if action == "url":
            url = (data.get("buttonUrl") or "").strip()
            if url:
                return InlineKeyboardButton(text=text, url=url)
        elif action == "web_app":
            url = (data.get("buttonWebAppUrl") or "").strip()
            if url:
                return InlineKeyboardButton(text=text, web_app=WebAppInfo(url=url))
        elif action == "copy":
            copy_text = (data.get("buttonCopyText") or "").strip()
            if copy_text:
                return InlineKeyboardButton(text=text, copy_text=CopyTextButton(text=copy_text))
        return InlineKeyboardButton(text=text, callback_data=f"btn:{btn.get('id') or ''}"[:64])

    def build_reply_button(btn: dict) -> Optional[KeyboardButton]:
        data = btn.get("data", {})
        text = (data.get("buttonText") or data.get("label") or "").strip()
        if not text:
            return None
        action = (data.get("replyAction") or "text").strip()
        if action == "web_app":
            url = (data.get("replyWebAppUrl") or "").strip()
            if url:
                return KeyboardButton(text=text, web_app=WebAppInfo(url=url))
        return KeyboardButton(text=text)

    if inline_rows:
        return InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    inline_btn
                    for btn in row
                    if (inline_btn := build_inline_button(btn)) is not None
                ]
                for row in inline_rows
            ]
        )

    if has_clear:
        return ReplyKeyboardRemove(remove_keyboard=True)

    if reply_rows:
        return ReplyKeyboardMarkup(
            keyboard=[
                [
                    reply_btn
                    for btn in row
                    if (reply_btn := build_reply_button(btn)) is not None
                ]
                for row in reply_rows
            ],
            resize_keyboard=True,
        )

    return None


@app.post("/uploads/images")
async def upload_images(files: List[UploadFile] = File(...)) -> dict:
    urls: List[str] = []
    for upload in files:
        ext = os.path.splitext(upload.filename or "")[1].lower() or ".jpg"
        name = f"{uuid.uuid4().hex}{ext}"
        path = os.path.join(UPLOAD_DIR, name)
        content = await upload.read()
        with open(path, "wb") as file_obj:
            file_obj.write(content)
        urls.append(f"/uploads/{name}")
    return {"urls": urls}


app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")


def collect_media_urls(flow: Flow, source_id: str, kind: str) -> List[str]:
    field_map = {
        "image": "imageUrls",
        "video": "videoUrls",
        "audio": "audioUrls",
        "document": "documentUrls",
    }
    field = field_map.get(kind)
    if not field:
        return []
    nodes_by_id = {node.get("id"): node for node in flow.nodes}
    urls: List[str] = []
    for edge in flow.edges:
        if edge.get("source") != source_id:
            continue
        target_node = nodes_by_id.get(edge.get("target"))
        if not target_node:
            continue
        target_data = target_node.get("data", {})
        if target_data.get("kind") != kind:
            continue
        items = target_data.get(field) or []
        for item in items:
            if isinstance(item, str) and item.strip():
                urls.append(item.strip())
    return urls


def collect_image_urls(flow: Flow, source_id: str) -> List[str]:
    return collect_media_urls(flow, source_id, "image")


def collect_video_urls(flow: Flow, source_id: str) -> List[str]:
    return collect_media_urls(flow, source_id, "video")


def collect_audio_urls(flow: Flow, source_id: str) -> List[str]:
    return collect_media_urls(flow, source_id, "audio")


def collect_document_urls(flow: Flow, source_id: str) -> List[str]:
    return collect_media_urls(flow, source_id, "document")


def resolve_upload_source(url: str):
    if not isinstance(url, str):
        return None
    cleaned = url.strip()
    if not cleaned:
        return None
    if cleaned.startswith("/uploads/"):
        filename = os.path.basename(cleaned)
        path = os.path.join(UPLOAD_DIR, filename)
        if os.path.exists(path):
            return FSInputFile(path)
        return cleaned
    if cleaned.startswith("http://") or cleaned.startswith("https://"):
        parsed = urlparse(cleaned)
        if parsed.hostname in ("localhost", "127.0.0.1"):
            filename = os.path.basename(parsed.path)
            path = os.path.join(UPLOAD_DIR, filename)
            if os.path.exists(path):
                return FSInputFile(path)
        return cleaned
    path = os.path.join(UPLOAD_DIR, os.path.basename(cleaned))
    if os.path.exists(path):
        return FSInputFile(path)
    return cleaned


async def send_images(
    message: Message,
    urls: List[str],
    caption: str = "",
    reply_markup=None,
    target_chat_id: Optional[int] = None,
) -> bool:
    sources = [resolve_upload_source(url) for url in urls]
    sources = [source for source in sources if source]
    if not sources:
        return False
    chat_id = target_chat_id if target_chat_id is not None else message.chat.id
    use_bot = target_chat_id is not None and message.chat and target_chat_id != message.chat.id
    if len(sources) == 1:
        if use_bot:
            await message.bot.send_photo(chat_id, sources[0], caption=caption or None, reply_markup=reply_markup)
        else:
            await message.answer_photo(sources[0], caption=caption or None, reply_markup=reply_markup)
        return True
    media = []
    for idx, source in enumerate(sources):
        if idx == 0 and caption:
            media.append(InputMediaPhoto(media=source, caption=caption))
        else:
            media.append(InputMediaPhoto(media=source))
    if use_bot:
        await message.bot.send_media_group(chat_id, media)
    else:
        await message.answer_media_group(media)
    return True


async def send_videos(
    message: Message,
    urls: List[str],
    caption: str = "",
    reply_markup=None,
    target_chat_id: Optional[int] = None,
) -> bool:
    sources = [resolve_upload_source(url) for url in urls]
    sources = [source for source in sources if source]
    if not sources:
        return False
    chat_id = target_chat_id if target_chat_id is not None else message.chat.id
    use_bot = target_chat_id is not None and message.chat and target_chat_id != message.chat.id
    if len(sources) == 1:
        if use_bot:
            await message.bot.send_video(chat_id, sources[0], caption=caption or None, reply_markup=reply_markup)
        else:
            await message.answer_video(sources[0], caption=caption or None, reply_markup=reply_markup)
        return True
    media = []
    for idx, source in enumerate(sources):
        if idx == 0 and caption:
            media.append(InputMediaVideo(media=source, caption=caption))
        else:
            media.append(InputMediaVideo(media=source))
    if use_bot:
        await message.bot.send_media_group(chat_id, media)
    else:
        await message.answer_media_group(media)
    return True


async def send_audios(
    message: Message,
    urls: List[str],
    caption: str = "",
    reply_markup=None,
    target_chat_id: Optional[int] = None,
) -> bool:
    sources = [resolve_upload_source(url) for url in urls]
    sources = [source for source in sources if source]
    if not sources:
        return False
    chat_id = target_chat_id if target_chat_id is not None else message.chat.id
    use_bot = target_chat_id is not None and message.chat and target_chat_id != message.chat.id
    if len(sources) == 1:
        if use_bot:
            await message.bot.send_audio(chat_id, sources[0], caption=caption or None, reply_markup=reply_markup)
        else:
            await message.answer_audio(sources[0], caption=caption or None, reply_markup=reply_markup)
        return True
    media = []
    for idx, source in enumerate(sources):
        if idx == 0 and caption:
            media.append(InputMediaAudio(media=source, caption=caption))
        else:
            media.append(InputMediaAudio(media=source))
    if use_bot:
        await message.bot.send_media_group(chat_id, media)
    else:
        await message.answer_media_group(media)
    return True


async def send_documents(
    message: Message,
    urls: List[str],
    caption: str = "",
    reply_markup=None,
    target_chat_id: Optional[int] = None,
) -> bool:
    sources = [resolve_upload_source(url) for url in urls]
    sources = [source for source in sources if source]
    if not sources:
        return False
    chat_id = target_chat_id if target_chat_id is not None else message.chat.id
    use_bot = target_chat_id is not None and message.chat and target_chat_id != message.chat.id
    if len(sources) == 1:
        if use_bot:
            await message.bot.send_document(chat_id, sources[0], caption=caption or None, reply_markup=reply_markup)
        else:
            await message.answer_document(sources[0], caption=caption or None, reply_markup=reply_markup)
        return True
    media = []
    for idx, source in enumerate(sources):
        if idx == 0 and caption:
            media.append(InputMediaDocument(media=source, caption=caption))
        else:
            media.append(InputMediaDocument(media=source))
    if use_bot:
        await message.bot.send_media_group(chat_id, media)
    else:
        await message.answer_media_group(media)
    return True


async def send_content_node(
    flow: Flow,
    message: Message,
    target_node: dict,
    target_chat_id: Optional[int] = None,
    source_user: Optional[object] = None,
    row_data: Optional[dict] = None,
) -> None:
    payload = target_node.get("data", {})
    kind = payload.get("kind")
    if kind == "delete_message":
        try:
            await message.delete()
        except Exception:
            pass
        return
    is_edit = kind == "edit_message"
    if kind == "message" or is_edit:
        if is_edit:
            try:
                await message.delete()
            except Exception:
                pass
        raw_text = payload.get("editMessageText") if is_edit else payload.get("messageText")
        message_text = render_template((raw_text or "").strip(), message, source_user, row_data=row_data)
        image_urls = collect_image_urls(flow, target_node.get("id") or "")
        video_urls = collect_video_urls(flow, target_node.get("id") or "")
        audio_urls = collect_audio_urls(flow, target_node.get("id") or "")
        document_urls = collect_document_urls(flow, target_node.get("id") or "")
        reply_markup = build_reply_markup(flow, target_node.get("id") or "")
        caption_used = False
        reply_used = False
        if image_urls:
            caption = message_text if message_text and not caption_used else ""
            await send_images(
                message,
                image_urls,
                caption=caption,
                reply_markup=reply_markup if not reply_used else None,
                target_chat_id=target_chat_id,
            )
            if caption:
                caption_used = True
                reply_used = True
        if video_urls:
            caption = message_text if message_text and not caption_used else ""
            await send_videos(
                message,
                video_urls,
                caption=caption,
                reply_markup=reply_markup if not reply_used else None,
                target_chat_id=target_chat_id,
            )
            if caption:
                caption_used = True
                reply_used = True
        if audio_urls:
            caption = message_text if message_text and not caption_used else ""
            await send_audios(
                message,
                audio_urls,
                caption=caption,
                reply_markup=reply_markup if not reply_used else None,
                target_chat_id=target_chat_id,
            )
            if caption:
                caption_used = True
                reply_used = True
        if document_urls:
            caption = message_text if message_text and not caption_used else ""
            await send_documents(
                message,
                document_urls,
                caption=caption,
                reply_markup=reply_markup if not reply_used else None,
                target_chat_id=target_chat_id,
            )
            if caption:
                caption_used = True
                reply_used = True
        if not (image_urls or video_urls or audio_urls or document_urls):
            if message_text:
                if target_chat_id is not None and message.chat and target_chat_id != message.chat.id:
                    await message.bot.send_message(target_chat_id, message_text, reply_markup=reply_markup)
                else:
                    await message.answer(message_text, reply_markup=reply_markup)
        return
    if kind == "image":
        urls = payload.get("imageUrls") or []
        reply_markup = build_reply_markup(flow, target_node.get("id") or "")
        await send_images(message, urls, reply_markup=reply_markup, target_chat_id=target_chat_id)
        return
    if kind == "video":
        urls = payload.get("videoUrls") or []
        reply_markup = build_reply_markup(flow, target_node.get("id") or "")
        await send_videos(message, urls, reply_markup=reply_markup, target_chat_id=target_chat_id)
        return
    if kind == "audio":
        urls = payload.get("audioUrls") or []
        reply_markup = build_reply_markup(flow, target_node.get("id") or "")
        await send_audios(message, urls, reply_markup=reply_markup, target_chat_id=target_chat_id)
        return
    if kind == "document":
        urls = payload.get("documentUrls") or []
        reply_markup = build_reply_markup(flow, target_node.get("id") or "")
        await send_documents(message, urls, reply_markup=reply_markup, target_chat_id=target_chat_id)
        return
    message_text = render_template((payload.get("messageText") or "").strip(), message, source_user, row_data=row_data)
    image_urls = collect_image_urls(flow, target_node.get("id") or "")
    video_urls = collect_video_urls(flow, target_node.get("id") or "")
    audio_urls = collect_audio_urls(flow, target_node.get("id") or "")
    document_urls = collect_document_urls(flow, target_node.get("id") or "")
    reply_markup = build_reply_markup(flow, target_node.get("id") or "")
    caption_used = False
    reply_used = False
    if image_urls:
        caption = message_text if message_text and not caption_used else ""
        await send_images(
            message,
            image_urls,
            caption=caption,
            reply_markup=reply_markup if not reply_used else None,
            target_chat_id=target_chat_id,
        )
        if caption:
            caption_used = True
            reply_used = True
    if video_urls:
        caption = message_text if message_text and not caption_used else ""
        await send_videos(
            message,
            video_urls,
            caption=caption,
            reply_markup=reply_markup if not reply_used else None,
            target_chat_id=target_chat_id,
        )
        if caption:
            caption_used = True
            reply_used = True
    if audio_urls:
        caption = message_text if message_text and not caption_used else ""
        await send_audios(
            message,
            audio_urls,
            caption=caption,
            reply_markup=reply_markup if not reply_used else None,
            target_chat_id=target_chat_id,
        )
        if caption:
            caption_used = True
            reply_used = True
    if document_urls:
        caption = message_text if message_text and not caption_used else ""
        await send_documents(
            message,
            document_urls,
            caption=caption,
            reply_markup=reply_markup if not reply_used else None,
            target_chat_id=target_chat_id,
        )
        if caption:
            caption_used = True
            reply_used = True
    if not (image_urls or video_urls or audio_urls or document_urls):
        if message_text:
            if target_chat_id is not None and message.chat and target_chat_id != message.chat.id:
                await message.bot.send_message(target_chat_id, message_text, reply_markup=reply_markup)
            else:
                await message.answer(message_text, reply_markup=reply_markup)


async def send_targets_with_delay(
    flow: Flow,
    message: Message,
    targets: List[Tuple[dict, float, Optional[int], Optional[dict]]],
    source_user: Optional[object] = None,
) -> None:
    if not targets:
        return
    ordered = sorted(targets, key=lambda item: item[1])
    elapsed = 0.0
    for target_node, delay, target_chat_id, row_data in ordered:
        wait_time = delay - elapsed
        if wait_time > 0:
            await asyncio.sleep(wait_time)
            elapsed = delay
        await send_content_node(flow, message, target_node, target_chat_id, source_user, row_data)


async def run_bot_polling(bot: Bot) -> None:
    if not bot.token:
        return
    dispatcher = Dispatcher()
    telegram_bot = TelegramBot(bot.token)
    bot_user = await telegram_bot.get_me()
    bot_user_id = bot_user.id if bot_user else None
    chat_admin_cache: Dict[int, bool] = {}
    last_run_by_task: Dict[str, datetime] = {}
    stop_event = asyncio.Event()
    scheduler_task = asyncio.create_task(task_scheduler_loop(bot, telegram_bot, stop_event, last_run_by_task))
    async def handler(message: Message) -> None:
        await ensure_user_row(bot.id, message.from_user, telegram_bot)
        if bot_user_id is not None:
            await ensure_chat_row(bot.id, message.chat, telegram_bot, bot_user_id, chat_admin_cache)
        command = normalize_command(message.text or "")
        flow = FLOW_CACHE.get(bot.id) or bot.flow
        user_id = message.from_user.id if message.from_user else None
        if command:
            command_node = find_command_node(flow, command)
            if command_node:
                targets = await collect_content_targets_with_delay(
                    flow, command_node.get("id") or "", message, bot.id, user_id
                )
                await send_targets_with_delay(flow, message, targets, message.from_user)
                return
        reply_button = find_reply_button_by_text(flow, message.text or "")
        if reply_button:
            targets = await collect_content_targets_with_delay(
                flow, reply_button.get("id") or "", message, bot.id, user_id
            )
            if targets:
                await send_targets_with_delay(flow, message, targets, message.from_user)
                return
        webhook_nodes = [node for node in flow.nodes if node.get("data", {}).get("kind") == "webhook"]
        if webhook_nodes:
            all_targets: List[Tuple[dict, float, Optional[int], Optional[dict]]] = []
            for webhook_node in webhook_nodes:
                all_targets.extend(
                    await collect_content_targets_with_delay(
                        flow, webhook_node.get("id") or "", message, bot.id, user_id
                    )
                )
            await send_targets_with_delay(flow, message, all_targets, message.from_user)

    dispatcher.message()(handler)

    async def channel_post_handler(message: Message) -> None:
        if bot_user_id is None:
            return
        await ensure_chat_row(bot.id, message.chat, telegram_bot, bot_user_id, chat_admin_cache)

    dispatcher.channel_post()(channel_post_handler)

    async def callback_handler(query: CallbackQuery) -> None:
        await ensure_user_row(bot.id, query.from_user, telegram_bot)
        if bot_user_id is not None and query.message:
            await ensure_chat_row(bot.id, query.message.chat, telegram_bot, bot_user_id, chat_admin_cache)
        data = (query.data or "").strip()
        await query.answer()
        if not data:
            return
        if not query.message:
            return
        flow = FLOW_CACHE.get(bot.id) or bot.flow
        user_id = query.from_user.id if query.from_user else None
        if data.startswith("btn:"):
            button_id = data[4:]
            targets = await collect_content_targets_with_delay(flow, button_id, query.message, bot.id, user_id)
            await send_targets_with_delay(flow, query.message, targets, query.from_user)
            return
        if data.startswith("/"):
            flow = FLOW_CACHE.get(bot.id) or bot.flow
            command = normalize_command(data)
            if command:
                command_node = find_command_node(flow, command)
                if command_node:
                    targets = await collect_content_targets_with_delay(
                        flow, command_node.get("id") or "", query.message, bot.id, user_id
                    )
                    await send_targets_with_delay(flow, query.message, targets, query.from_user)
                return
        await query.message.answer(data)

    dispatcher.callback_query()(callback_handler)

    RUNNING_BOTS[bot.id] = {"task": asyncio.current_task(), "stop": stop_event, "scheduler": scheduler_task}

    try:
        await telegram_bot.delete_webhook(drop_pending_updates=True)
        await dispatcher.start_polling(telegram_bot, stop_event=stop_event, handle_signals=False)
    finally:
        await telegram_bot.session.close()
        if scheduler_task:
            scheduler_task.cancel()
        RUNNING_BOTS.pop(bot.id, None)


async def stop_bot_task(bot_id: str) -> None:
    entry = RUNNING_BOTS.get(bot_id)
    if not entry:
        return
    stop_event = entry.get("stop")
    task = entry.get("task")
    scheduler = entry.get("scheduler")
    if isinstance(stop_event, asyncio.Event):
        stop_event.set()
    if isinstance(task, asyncio.Task):
        try:
            await asyncio.wait_for(task, timeout=5)
        except asyncio.TimeoutError:
            task.cancel()
        except Exception:
            pass
    if isinstance(scheduler, asyncio.Task):
        try:
            await asyncio.wait_for(scheduler, timeout=5)
        except asyncio.TimeoutError:
            scheduler.cancel()
        except Exception:
            pass
    RUNNING_BOTS.pop(bot_id, None)


@app.get("/health")
def health() -> dict:
    return {"status": "ok"}


@app.post("/bots", response_model=Bot)
def create_bot(payload: BotCreate) -> Bot:
    bot_id = str(uuid.uuid4())
    flow = Flow().model_dump()
    with get_connection() as conn:
        conn.execute(
            "INSERT INTO bots (id, name, token, status, flow) VALUES (?, ?, ?, ?, ?)",
            (bot_id, payload.name, payload.token, "stopped", json.dumps(flow)),
        )
        conn.commit()
    return get_bot_or_404(bot_id)


@app.get("/bots", response_model=List[Bot])
def list_bots() -> List[Bot]:
    with get_connection() as conn:
        rows = conn.execute("SELECT * FROM bots ORDER BY rowid DESC").fetchall()
    return [row_to_bot(row) for row in rows]


@app.get("/bots/{bot_id}", response_model=Bot)
def get_bot(bot_id: str) -> Bot:
    return get_bot_or_404(bot_id)


@app.get("/bots/{bot_id}/users")
def list_bot_users(bot_id: str) -> List[dict]:
    get_bot_or_404(bot_id)
    return list_users(bot_id)


@app.get("/bots/{bot_id}/chats")
def list_bot_chats(bot_id: str) -> List[dict]:
    get_bot_or_404(bot_id)
    return list_chats(bot_id)


@app.get("/bots/{bot_id}/files/excel/{name}")
def download_excel_file(bot_id: str, name: str) -> FileResponse:
    get_bot_or_404(bot_id)
    path = get_excel_path(bot_id, name)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(path, media_type="text/csv", filename=os.path.basename(path))


@app.post("/bots/{bot_id}/files/excel/upload")
async def upload_excel_file(bot_id: str, file: UploadFile = File(...)) -> dict:
    get_bot_or_404(bot_id)
    original = file.filename or "data.csv"
    base_name, ext = os.path.splitext(original)
    base_name = base_name or "data"
    safe_name = sanitize_filename(base_name)
    path = get_excel_path(bot_id, safe_name)
    ext = ext.lower()
    content = await file.read()
    if ext in ("", ".csv"):
        with open(path, "wb") as file_obj:
            file_obj.write(content)
        return {"name": safe_name}
    if ext in (".xlsx", ".xls"):
        try:
            import pandas as pd

            engine = "openpyxl" if ext == ".xlsx" else "xlrd"
            df = pd.read_excel(BytesIO(content), engine=engine)
            df.to_csv(path, index=False)
            return {"name": safe_name}
        except ImportError:
            pass
        except Exception:
            pass
        if ext == ".xlsx":
            try:
                from openpyxl import load_workbook
            except ImportError:
                raise HTTPException(
                    status_code=400,
                    detail="Install openpyxl/pandas to upload xlsx/xls files",
                )
            wb = load_workbook(BytesIO(content), data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            with open(path, "w", encoding="utf-8", newline="") as file_obj:
                writer = csv.writer(file_obj)
                if rows:
                    headers = [str(cell) if cell is not None else "" for cell in rows[0]]
                    writer.writerow(headers)
                    for row in rows[1:]:
                        writer.writerow(["" if cell is None else cell for cell in row])
            return {"name": safe_name}
        raise HTTPException(
            status_code=400,
            detail="Install pandas/xlrd to upload xls files",
        )
    raise HTTPException(status_code=400, detail="Unsupported file type")


@app.get("/bots/{bot_id}/files/text/{name}")
def download_text_file(bot_id: str, name: str) -> FileResponse:
    get_bot_or_404(bot_id)
    path = get_text_path(bot_id, name)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(path, media_type="text/plain", filename=os.path.basename(path))


@app.post("/bots/{bot_id}/files/text/upload")
async def upload_text_file(bot_id: str, file: UploadFile = File(...)) -> dict:
    get_bot_or_404(bot_id)
    original = file.filename or "data.txt"
    base_name = os.path.splitext(original)[0] or "data"
    safe_name = sanitize_filename(base_name)
    path = get_text_path(bot_id, safe_name)
    content = await file.read()
    with open(path, "wb") as file_obj:
        file_obj.write(content)
    return {"name": safe_name}


@app.patch("/bots/{bot_id}", response_model=Bot)
def update_bot(bot_id: str, payload: BotUpdate) -> Bot:
    bot = get_bot_or_404(bot_id)
    data = bot.model_dump()
    update = payload.model_dump(exclude_unset=True)
    data.update(update)
    updated = Bot(**data)
    update_bot_row(updated)
    return updated


@app.delete("/bots/{bot_id}")
def delete_bot(bot_id: str) -> dict:
    with get_connection() as conn:
        row = conn.execute("SELECT 1 FROM bots WHERE id = ?", (bot_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Bot not found")
        conn.execute("DELETE FROM bots WHERE id = ?", (bot_id,))
        conn.commit()
    asyncio.create_task(stop_bot_task(bot_id))
    return {"deleted": True}


@app.post("/bots/{bot_id}/flow", response_model=Bot)
async def save_flow(bot_id: str, flow: Flow) -> Bot:
    bot = get_bot_or_404(bot_id)
    updated = bot.model_copy(update={"flow": flow})
    update_bot_row(updated)
    FLOW_CACHE[bot_id] = flow
    return updated


@app.post("/bots/{bot_id}/start", response_model=Bot)
async def start_bot(bot_id: str) -> Bot:
    bot = get_bot_or_404(bot_id)
    if not bot.token:
        raise HTTPException(status_code=400, detail="Bot token is required")
    FLOW_CACHE[bot_id] = bot.flow
    if bot_id in RUNNING_BOTS:
        updated = bot.model_copy(update={"status": "running"})
        update_bot_row(updated)
        return updated
    await stop_bot_task(bot_id)
    updated = bot.model_copy(update={"status": "running"})
    update_bot_row(updated)
    asyncio.create_task(run_bot_polling(updated))
    return updated


@app.post("/bots/{bot_id}/stop", response_model=Bot)
async def stop_bot(bot_id: str) -> Bot:
    bot = get_bot_or_404(bot_id)
    await stop_bot_task(bot_id)
    updated = bot.model_copy(update={"status": "stopped"})
    update_bot_row(updated)
    return updated
